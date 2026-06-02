from datetime import datetime
from tkinter import filedialog

import customtkinter as ctk
from openpyxl import Workbook

from client import theme as T


class StatisticsPage(ctk.CTkFrame):
    SUMMARY_META = [
        ("Toplam Satış", T.PRIMARY),
        ("Toplam Gelir", T.SUCCESS),
        ("Toplam Yükleme", T.PURPLE),
    ]

    def __init__(self, master, app):
        super().__init__(master, fg_color="transparent")
        self.app = app
        self.stats = None

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        header = ctk.CTkFrame(self, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=T.PAD_PAGE, pady=(T.PAD_PAGE, 12))
        header.grid_columnconfigure(0, weight=1)

        title_block = T.build_page_header(header, "İstatistikler", "Satış, gelir ve müşteri raporları.")
        title_block.grid(row=0, column=0, sticky="w")

        btn_row = ctk.CTkFrame(header, fg_color="transparent")
        btn_row.grid(row=0, column=1, sticky="e")

        T.ghost_btn(btn_row, text="Yenile", width=100, command=self._load).pack(side="left", padx=(0, 10))
        ctk.CTkButton(
            btn_row,
            text="Excel İndir",
            width=130,
            height=40,
            corner_radius=T.RADIUS_SM,
            fg_color=T.PURPLE,
            hover_color=T.PURPLE_HOVER,
            font=T.font_body(13),
            command=self._export_excel,
        ).pack(side="left")

        self.summary = ctk.CTkFrame(self, fg_color="transparent")
        self.summary.grid(row=1, column=0, sticky="ew", padx=T.PAD_PAGE, pady=12)
        for i in range(3):
            self.summary.grid_columnconfigure(i, weight=1)

        self.summary_labels = []
        for i, (title, accent) in enumerate(self.SUMMARY_META):
            card = T.card(self.summary)
            card.grid(row=0, column=i, padx=8, pady=0, sticky="nsew")

            stripe = ctk.CTkFrame(card, height=3, fg_color=accent, corner_radius=0)
            stripe.pack(fill="x")

            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(fill="x", padx=20, pady=18)

            ctk.CTkLabel(inner, text=title, font=T.font_body(13), text_color=T.TEXT_MUTED).pack(anchor="w")
            lbl = ctk.CTkLabel(inner, text="—", font=T.font_stat(), text_color=T.TEXT)
            lbl.pack(anchor="w", pady=(6, 0))
            self.summary_labels.append(lbl)

        self.tabs = ctk.CTkTabview(
            self,
            fg_color=T.SURFACE,
            segmented_button_fg_color=T.SURFACE_ALT,
            segmented_button_selected_color=T.PRIMARY,
            segmented_button_selected_hover_color=T.PRIMARY_HOVER,
            segmented_button_unselected_color=T.SURFACE_ALT,
            segmented_button_unselected_hover_color=T.BORDER,
            corner_radius=T.RADIUS_LG,
            border_width=1,
            border_color=T.BORDER,
        )
        self.tabs.grid(row=2, column=0, sticky="nsew", padx=T.PAD_PAGE, pady=(0, T.PAD_PAGE))
        self.tabs.add("Ürün İstatistikleri")
        self.tabs.add("Müşteri Alışverişleri")

        self.products_table = ctk.CTkScrollableFrame(
            self.tabs.tab("Ürün İstatistikleri"),
            fg_color="transparent",
        )
        self.products_table.pack(fill="both", expand=True, padx=12, pady=12)

        self.customer_table = ctk.CTkScrollableFrame(
            self.tabs.tab("Müşteri Alışverişleri"),
            fg_color="transparent",
        )
        self.customer_table.pack(fill="both", expand=True, padx=12, pady=12)

    def on_show(self):
        self._load()

    def _load(self):
        for lbl in self.summary_labels:
            lbl.configure(text="...")

        def work():
            return self.app.api.get_statistics()

        def on_success(stats):
            self.stats = stats
            self.summary_labels[0].configure(text=str(stats["total_sales_count"]))
            self.summary_labels[1].configure(text=f"{stats['total_revenue']:.2f} ₺")
            self.summary_labels[2].configure(text=f"{stats['total_topups']:.2f} ₺")
            self._render_products(stats["products"])
            self._render_customers(stats["customer_purchases"])

        def on_error(exc):
            for lbl in self.summary_labels:
                lbl.configure(text="—")
            self.app.show_error(str(exc))

        self.app.run_in_background(work, on_success=on_success, on_error=on_error)

    def _render_products(self, products: list):
        for w in self.products_table.winfo_children():
            w.destroy()

        headers = ["Ürün ID", "Ad", "Fiyat", "Stok", "Satılan", "Gelir"]
        self._table_header(self.products_table, headers)

        for p in products:
            self._table_row(
                self.products_table,
                [
                    p["product_code"],
                    p["name"],
                    f"{p['price']:.2f}",
                    str(p["stock"]),
                    str(p["total_sold"]),
                    f"{p['total_revenue']:.2f}",
                ],
            )

    def _render_customers(self, rows: list):
        for w in self.customer_table.winfo_children():
            w.destroy()

        headers = ["Müşteri", "Kart UID", "Ürün", "Adet", "Harcama"]
        self._table_header(self.customer_table, headers)

        if not rows:
            ctk.CTkLabel(
                self.customer_table,
                text="Henüz satış kaydı yok.",
                font=T.font_body(13),
                text_color=T.TEXT_MUTED,
            ).pack(anchor="w", padx=8, pady=12)
            return

        for r in rows:
            self._table_row(
                self.customer_table,
                [
                    r["customer_name"],
                    r["card_uid"],
                    f"{r['product_name']} ({r['product_code']})",
                    str(r["quantity"]),
                    f"{r['total_spent']:.2f}",
                ],
            )

    def _table_header(self, parent, headers: list):
        row = ctk.CTkFrame(parent, fg_color=T.SURFACE_ALT, corner_radius=T.RADIUS_SM)
        row.pack(fill="x", pady=(0, 6))
        widths = [100, 180, 80, 70, 80, 90]
        for h, w in zip(headers, widths):
            ctk.CTkLabel(
                row,
                text=h,
                width=w,
                anchor="w",
                font=T.font_body(12),
                text_color=T.TEXT_MUTED,
            ).pack(side="left", padx=8, pady=10)

    def _table_row(self, parent, values: list):
        row = ctk.CTkFrame(parent, fg_color=T.SURFACE, corner_radius=T.RADIUS_SM)
        row.pack(fill="x", pady=2)
        widths = [100, 180, 80, 70, 80, 90]
        for val, w in zip(values, widths):
            ctk.CTkLabel(
                row,
                text=str(val),
                width=w,
                anchor="w",
                font=T.font_body(12),
                text_color=T.TEXT,
            ).pack(side="left", padx=8, pady=8)

    def _export_excel(self):
        if not self.stats:
            self.app.show_error("Önce istatistikleri yükleyin.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"istatistikler_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        )
        if not path:
            return

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Urun Istatistikleri"
        ws1.append(["Urun ID", "Ad", "Fiyat", "Stok", "Satilan", "Gelir"])
        for p in self.stats["products"]:
            ws1.append(
                [
                    p["product_code"],
                    p["name"],
                    p["price"],
                    p["stock"],
                    p["total_sold"],
                    p["total_revenue"],
                ]
            )

        ws2 = wb.create_sheet("Musteri Alisverisleri")
        ws2.append(["Musteri", "Kart UID", "Urun Kodu", "Urun Adi", "Adet", "Harcama"])
        for r in self.stats["customer_purchases"]:
            ws2.append(
                [
                    r["customer_name"],
                    r["card_uid"],
                    r["product_code"],
                    r["product_name"],
                    r["quantity"],
                    r["total_spent"],
                ]
            )

        ws3 = wb.create_sheet("Ozet")
        ws3.append(["Metrik", "Deger"])
        ws3.append(["Toplam Satis Sayisi", self.stats["total_sales_count"]])
        ws3.append(["Toplam Gelir", self.stats["total_revenue"]])
        ws3.append(["Toplam Yukleme", self.stats["total_topups"]])

        wb.save(path)
        self.app.show_info(f"İstatistikler kaydedildi:\n{path}")
